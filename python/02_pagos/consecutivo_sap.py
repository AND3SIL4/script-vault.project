import pandas as pd  # type: ignore
import numpy as np  # type: ignore
from typing import Optional
from openpyxl import load_workbook  # type: ignore


class Consecutivo:
    """Clase para manejar la información de coaseguros"""

    def __init__(
        self,
        file_path: str,
        sheet_name: str,
        exception_file: str,
        consecutivo_sap_file: str,
        consecutivo_sheet: str
    ):
        self.path_file = file_path
        self.sheet_name = sheet_name
        self.exception_file = exception_file
        self.consecutivo_sap_file = consecutivo_sap_file
        self.consecutivo_sheet = consecutivo_sheet

    def read_excel(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """Method for returning a data frame"""
        return pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    def filter_file(
        self, data_frame: pd.DataFrame, cut_off_date: str, col_idx: int
    ) -> pd.DataFrame:
        """Method to filter the file according to the defined criteria"""

        # Parse the cutoff date to get the month and year
        date: pd.Timestamp = pd.to_datetime(
            cut_off_date, format="%d/%m/%Y", errors="coerce"
        )
        year = date.year
        month = date.month

        # Get the column name using the index position (col_idx)
        col_name = data_frame.columns[col_idx]

        # Convert the specified column to datetime type and assign it back to the DataFrame
        data_frame[col_name] = pd.to_datetime(data_frame[col_name], errors="coerce")

        # Filter DataFrame by month and year
        filtered_df: pd.DataFrame = data_frame[
            (data_frame[col_name].dt.month == month)
            & (data_frame[col_name].dt.year == year)
        ]

        return filtered_df  # Return the filtered DataFrame

    def update_data(
        self, consecutivo_inicial: int, consecutivo_final: int, lista_consecutivos: list
    ) -> None:
        try:

            # Cargar el archivo existente
            book = load_workbook(self.exception_file)
            sheet = book["CONSECUTIVO SAP"]

            # Limpiar los datos antiguos en la columna 1 (consecutivos pendientes)
            max_row = sheet.max_row  # Obtiene el número máximo de filas usadas
            for row in range(2, max_row + 1):  # Empieza en la fila 2
                sheet.cell(row=row, column=5).value = None  # Borra los valores antiguos

            # Actualizar valores
            sheet.cell(row=2, column=6).value = (
                consecutivo_inicial  # Consecutivo inicial
            )
            sheet.cell(row=2, column=7).value = consecutivo_final  # Consecutivo final

            # Actualizar lista de consecutivos pendientes
            row = 1
            for consecutivo in lista_consecutivos:
                row += 1
                sheet.cell(row=row, column=5).value = consecutivo

            # Guardar cambios
            book.save(self.exception_file)
            return True
        except Exception as e:
            print(f"ERROR {e}")
            return False

    def consecutivo(self, cut_off_date: str) -> str:
        # Pagos data frame
        pagos_file: pd.DataFrame = self.read_excel(self.path_file, self.sheet_name)
        # Consecutivo data frame
        consecutivo_file: pd.DataFrame = self.read_excel(
            self.consecutivo_sap_file, self.consecutivo_sheet
        )
        # Information from EXCEPTION FILE
        list_df: pd.DataFrame = self.read_excel(self.exception_file, "CONSECUTIVO SAP")
        # Consecutivo data frame after being filtered
        consecutivo_df = self.filter_file(consecutivo_file, cut_off_date, 0)
        # Pagos data frame after being filtered
        pagos_df = self.filter_file(pagos_file, cut_off_date, 72)

        # Initial variables (consecutivo final, consecutivos faltantes)
        final_consecutivo: int = int(list_df.iloc[0, 2])

        lista_drive: list[int] = consecutivo_df.iloc[:, 1].astype(int).to_list()
        # Remove the final consecutivo if it's in the drive list
        if final_consecutivo in lista_drive:
            lista_drive.remove(final_consecutivo)

        # List of values from the EXCEPTION FILE TODO: Validate first
        pending_list: list[int] = list_df.iloc[:, 0].dropna().astype(int).to_list()

        # List of values from PAGOS FILE without duplicates
        consecutivos_pagos: list = (
            pagos_df.iloc[:, 73].drop_duplicates().astype(int).to_list()
        )

        # 1. Check the pending list before make validation
        # Create a copy of data frames in order to do not affect the origins
        consecutivos_pagos_copy = consecutivos_pagos.copy()
        pending_list_copy = pending_list.copy()

        pending_list = [
            consecutivo
            for consecutivo in pending_list
            if consecutivo not in consecutivos_pagos_copy
        ]
        consecutivos_pagos = [
            consecutivo
            for consecutivo in consecutivos_pagos
            if consecutivo not in pending_list_copy
        ]

        # Size of the total different values after deleting matching values
        # length: int = len(consecutivo_df) - len(pending_list_copy)

        # Get the list with autoincrement (+1) starts from before final consecutivo
        consecutivos_to_validate: list[int] = []

        # # Fill up the list with autoincrement (+1) starts from before final consecutivo
        # for consecutivo in range(length):
        #     final_consecutivo += 1
        #     consecutivos_to_validate.append(final_consecutivo)

        while True:
            # Check the final consecutivo to validate
            if lista_drive[-1] == final_consecutivo:
                break
            final_consecutivo += 1
            consecutivos_to_validate.append(final_consecutivo)

        missing_consecutivos = [
            value
            for value in consecutivos_to_validate
            if value not in consecutivos_pagos
        ]

        # Create a data frame with consecutivos missing
        missing_consecutivos_df: pd.DataFrame = pd.DataFrame(
            missing_consecutivos, columns=["CONSECUTIVOS MISSING"]
        )
        # Merge files to get all data
        missing_consecutivos_df: pd.DataFrame = missing_consecutivos_df.merge(
            consecutivo_df,
            left_on=missing_consecutivos_df.columns[0],
            right_on=consecutivo_df.columns[1],
            how="left",
        )

        # If the records are associated to RED ASISTENCIAL i must'nt take in count
        final_df: pd.DataFrame = missing_consecutivos_df[
            missing_consecutivos_df.iloc[:, 3] != "RED ASISTENCIAL"
        ]

        # Create a list to write into the exception file
        append_list = pending_list + final_df.iloc[:, 0].dropna().to_list()

        print("Values pending to append in inconsistencies list:", append_list)

        data_updated: bool = self.update_data(
            consecutivos_to_validate[0], consecutivos_to_validate[-1], append_list
        )
        # Save the information into the file
        return (
            (
                ("Archivo excepciones actualizado correctamente")
                if data_updated
                else ("Error al actualizar el archivo de excepciones")
            ),
            data_updated,
        )


# * INITIALIZE THE VARIABLE TO INSTANCE THE MAIN CLASS
consecutivo: Optional[Consecutivo] = None


# * CALL THE MAIN FUNCTION WITH THE MAIN PARAMS
def main(params: dict) -> bool:
    try:
        global consecutivo

        # Get the variables
        file_path: str = params.get("file_path")
        sheet_name: str = params.get("sheet_name")
        exception_file: str = params.get("exception_file")
        consecutivo_sap_file: str = params.get("consecutivo_sap_file")
        consecutivo_sheet: str = params.get("consecutivo_sheet")

        # Pass the values to the constructor in the main class
        consecutivo = Consecutivo(
            file_path,
            sheet_name,
            exception_file,
            consecutivo_sap_file,
            consecutivo_sheet
        )
        return True
    except Exception as e:
        return f"ERROR: {e}"


def validate_consecutivo_sap(params: dict) -> str:
    try:
        # Set local variables
        cut_off_date: str = params.get("cut_off_date")
        validation: str = consecutivo.consecutivo(cut_off_date)
        return validation
    except Exception as e:
        return f"ERROR: {e}"


if __name__ == "__main__":
    params = {
        "file_path": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\TempFolder\BASE DE PAGOS.xlsx",
        "sheet_name": "PAGOS",
        "exception_file": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\InputFolder\EXCEPCIONES BASE PAGOS.xlsx",
        "consecutivo_sap_file": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\InputFolder\CONSECUTIVO SAP 2023.xlsx",
        "consecutivo_sheet" : "NUMERO DE PAGOO"
    }
    print(main(params))
    params = {"cut_off_date": "29/11/2024"}
    print(validate_consecutivo_sap(params))
