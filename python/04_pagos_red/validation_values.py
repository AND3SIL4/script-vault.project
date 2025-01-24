import pandas as pd  # type: ignore
from typing import Optional
from datetime import datetime
from openpyxl import load_workbook  # type: ignore


class ValuesValidation:
    def __init__(
        self,
        file_path: str,
        inconsistencies_file: str,
        exception_file: str,
        sheet_name,
        file_name: str,
        previous_file: str,
        temp_file: str,
        historic_file: str,
    ):
        self.file_path = file_path
        self.inconsistencies_file = inconsistencies_file
        self.exception_file = exception_file
        self.sheet_name = sheet_name
        self.file_name = file_name
        self.previous_file = previous_file
        self.temp_file = temp_file
        self.historic_file = historic_file

    def read_excel(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """Method for returning a data frame"""
        return pd.read_excel(
            file_path, sheet_name=sheet_name, engine="openpyxl", dtype=str
        )

    def save_inconsistencies_file(self, df: pd.DataFrame, new_sheet: str) -> bool:
        """Method to save the inconsistencies data frame into the inconsistencies file"""
        try:
            with pd.ExcelFile(self.inconsistencies_file, engine="openpyxl") as xls:
                if new_sheet in xls.sheet_names:
                    existing = pd.read_excel(
                        xls, engine="openpyxl", sheet_name=new_sheet
                    )
                    df = pd.concat([existing, df], ignore_index=True)

            with pd.ExcelWriter(
                self.inconsistencies_file,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            ) as writer:
                df.to_excel(writer, sheet_name=new_sheet, index=False)
                return True
        except Exception as e:
            print(f"Error: {e}")
            return False

    def excel_col_name(self, number) -> str:
        """Method to convert (1-based) to Excel column name"""
        result = ""
        while number > 0:
            number, reminder = divmod(number - 1, 26)
            result = chr(65 + reminder) + result
        return result

    def validate_inconsistencies(
        self, df: pd.DataFrame, col_idx, sheet_name: str
    ) -> str:
        """Method to validate the inconsistencies before append in a inconsistencies file"""
        if not df.empty:
            df = df.copy()
            if isinstance(col_idx, int):
                df[f"COORDENADAS"] = df.apply(
                    lambda row: f"{self.excel_col_name(col_idx + 1)}{row.name + 2}",
                    axis=1,
                )
            else:
                for i in col_idx:
                    df[f"COORDENADAS_{i + 2}"] = df.apply(
                        lambda row: f"{self.excel_col_name(i+1)}{row.name + 2}",
                        axis=1,
                    )
            self.save_inconsistencies_file(df, sheet_name)
            return "SUCCESS: Inconsistencies guardadas correctamente"
        else:
            return "INFO: Validacion realizada, no se encontraron inconsistencias"

    def get_file_date(self) -> str:
        """Method to get the file date"""
        start: str = self.file_name.find("(") + 1
        end: str = self.file_name.find(")")

        # Extract the text
        if start > 0 and end > 0:
            file_date = self.file_name[start:end]
            file_date = file_date.replace("-", "/")
            return file_date
        else:
            return "Error obteniendo fecha"

    def load_all_sheets(self) -> dict[str, pd.DataFrame]:
        """Method to load all the sheet into the 'Validador pagos' file
        and return a dictionary with the data frames for each sheet"""
        sheets = pd.read_excel(self.historic_file, sheet_name=None, engine="openpyxl")
        return sheets

    def ensure_sheet_exists(self, file_path: str, sheet_name: str) -> None:
        """Function to validate if the new sheet is in 'Validador pagos' file
        if the sheet does not exist, then it will be created"""
        try:
            workbook = load_workbook(file_path)
            # Verifica si la hoja ya existe
            if sheet_name not in workbook.sheetnames:
                # Si la hoja no existe, crea una nueva
                new_sheet = workbook.create_sheet(sheet_name)
                # Verifica si la hoja anterior existe
                if (
                    sheet_name != "Sheet1"
                    and f"{int(sheet_name)-1}" in workbook.sheetnames
                ):
                    previous_sheet = workbook[f"{int(sheet_name)-1}"]
                    # Copiar los encabezados de la hoja anterior
                    for col in range(1, previous_sheet.max_column + 1):
                        new_sheet.cell(
                            row=1,
                            column=col,
                            value=previous_sheet.cell(row=1, column=col).value,
                        )

                workbook.save(file_path)
                return True, f"Sheet {sheet_name} created successfully"
            return True, "Sheet is already exist"
        except Exception as e:
            return False, f"Error {e}"

    def validate_previous_counter(
        self, value: str, year: int, index: int, df_by_year: dict[str, pd.DataFrame]
    ) -> int:
        current_radicado_year = int(
            value[:4]
        )  # Convierte el año a entero para comparación
        if current_radicado_year != year:
            # Filtrar los DataFrames por los años relevantes
            relevant_years = range(current_radicado_year, year + 1)
            previous_dfs = []

            for y in relevant_years:
                # Solo accedemos a las hojas que están dentro del rango de años
                if str(y) in df_by_year:
                    # Filtramos las filas de la columna relevante (por índice) y convertimos a str
                    relevant_df = df_by_year[str(y)]
                    previous_dfs.append(relevant_df)

            entire_df: pd.DataFrame = pd.concat(previous_dfs, ignore_index=True)
            key_column_list: list[str] = (
                entire_df.iloc[:, index]
                .dropna()
                .astype(str)
                .str.replace(" -", "")
                .to_list()
            )
            counter: int = key_column_list.count(value)
            # print(f"{value} is: {counter} times year {current_radicado_year} - {year}")
            # Return the amount of values in the historical key list
            return counter
        return 0

    def save_inconsistencies(
        self,
        data_frame: pd.DataFrame,
        exception_sheet_name: str,
        exception_col_idx: int,
        validation_col_idx: int,
        col_idx_to_except: int,
        inconsistencies_sheet_name: str,
    ) -> str:
        valores_exception_df: pd.DataFrame = values_validation.read_excel(
            values_validation.exception_file, exception_sheet_name
        )
        valores_exception_list: list[str] = (
            valores_exception_df.iloc[:, exception_col_idx]
            .dropna()
            .astype(str)
            .to_list()
        )
        valores_inconsistencies: pd.DataFrame = data_frame[
            ~data_frame.iloc[:, validation_col_idx]
        ]
        valores_inconsistencies = valores_inconsistencies[
            ~valores_inconsistencies.iloc[:, col_idx_to_except].isin(
                valores_exception_list
            )
        ]
        # Save the inconsistencies
        values_validation.validate_inconsistencies(
            valores_inconsistencies,
            [col_idx_to_except, validation_col_idx],
            inconsistencies_sheet_name,
        )

    def save_inconsistencies_values(
        self,
        data_frame: pd.DataFrame,
        exception_sheet_name: str,
        exception_col: int,
        validation_col: int,
        list_col: int,
        inconsistencies_sheet_name: str,
    ) -> str:
        exception_df: pd.DataFrame = values_validation.read_excel(
            values_validation.exception_file, exception_sheet_name
        )
        radicados_exception_list: list[str] = (
            exception_df.iloc[:, exception_col].dropna().astype(str).to_list()
        )
        radicados_inconsistencies: pd.DataFrame = data_frame[
            data_frame.iloc[:, validation_col].astype(int) > 1
        ]
        radicados_inconsistencies = radicados_inconsistencies[
            ~radicados_inconsistencies.iloc[:, list_col].isin(radicados_exception_list)
        ]
        # Save the inconsistencies
        values_validation.validate_inconsistencies(
            radicados_inconsistencies,
            [list_col, validation_col],
            inconsistencies_sheet_name,
        )


# Instance the main class
values_validation: Optional[ValuesValidation] = None


def extract_data_from_propuesta(data_frame: pd.DataFrame) -> pd.DataFrame:
    """Method to get the importan data from propuesta de pagos file and the return it into a data frame"""
    # Filter data frame by number of index only with the important columns
    # N° radicado casa matriz = 2
    # Valor movimiento 100% = 45
    data_frame = data_frame.iloc[:, [2, 45]].copy()
    # Drop rows with NaN values in the first column (N° radicado casa matriz)
    data_frame.dropna(subset=[data_frame.columns[0]], inplace=True)
    # Add additional columns
    data_frame["LLAVE RADICADO + MOVIMIENTO"] = (
        data_frame[data_frame.columns[0]] + " " + data_frame[data_frame.columns[1]]
    )
    data_frame["AÑO"] = datetime.now().year
    data_frame.iloc[:, 0] = data_frame.iloc[:, 0]
    return data_frame


def cross_file(propuesta_df: pd.DataFrame, acm_df: pd.DataFrame) -> pd.DataFrame:
    """Method to cross the propuesta and acm report and return the merged data frame"""
    # Merge the two data frames based on the radicado number
    merged_df = pd.merge(
        propuesta_df,
        acm_df,
        left_on=propuesta_df.columns[0],
        right_on=acm_df.columns[0],
        how="left",
        suffixes=("_PROPUESTA", "_ACM"),
    )
    # Filter file before return it
    merged_df = merged_df[
        [
            "AÑO",
            "No DE RADICADO CASA MATRIZ",
            "VR. MOVIMIENTO 100%",
            "LLAVE RADICADO + MOVIMIENTO",
            "Valor Liquidado",
            "valor aprobado",
        ]
    ]
    return merged_df


def apply_formulas(data_frame: pd.DataFrame, historical_df: pd.DataFrame) -> None:
    """Method to apply formulas to the merged data frame"""
    # Convertir columnas a valores numéricos
    columns_to_convert = ["Valor Liquidado", "valor aprobado", "VR. MOVIMIENTO 100%"]
    data_frame[columns_to_convert] = (
        data_frame[columns_to_convert]
        .replace({",": "", ".": ""}, regex=True)
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
    )

    # Aplicar fórmulas
    data_frame[historical_df.columns[6]] = (
        data_frame["valor aprobado"] - data_frame["Valor Liquidado"]
    )

    # Validate valores reported
    data_frame[historical_df.columns[7]] = data_frame.iloc[:, 2].astype(
        int
    ) == data_frame.iloc[:, 5].astype(int)

    # Validate records duplicates
    ## Radicado column
    add_number_of_duplicates(
        data_frame=data_frame,
        historical_df=historical_df,
        index_col=1,
        historical_index_col=8,
    )

    ## Key column: (radicado number + concepto value)
    add_number_of_duplicates(
        data_frame=data_frame,
        historical_df=historical_df,
        index_col=3,
        historical_index_col=9,
    )

    # Validate format values in columns
    data_frame[historical_df.columns[10]] = data_frame[data_frame.columns[1]].apply(
        lambda value: str(value).replace(",", "").replace(".", "").isdigit()
    )
    data_frame[historical_df.columns[11]] = data_frame[data_frame.columns[2]].apply(
        lambda value: str(value).replace(",", "").replace(".", "").isdigit()
    )

    data_frame[historical_df.columns[12]] = values_validation.get_file_date()
    # Get the current year
    year = datetime.now().year

    # Validate if the new sheet exist and created
    sheet_created = values_validation.ensure_sheet_exists(
        file_path=values_validation.historic_file, sheet_name=str(year)
    )
    if not sheet_created:
        raise Exception("Error creating the sheet name into 'Validador Pagos' file")

    # Load all sheets only one time
    df_by_year = values_validation.load_all_sheets()

    # Get the radicado column
    valores_columna = data_frame.iloc[:, 1].astype(str)

    # Create a new column with the results
    data_frame["is_previous_radicado"] = [
        values_validation.validate_previous_counter(
            value=valor, year=year, index=1, df_by_year=df_by_year
        )
        for valor in valores_columna
    ]
    #  Get the key column
    valores_columna = data_frame.iloc[:, 3].astype(str)

    #  Create a new column with the results
    data_frame["is_previous_key"] = [
        values_validation.validate_previous_counter(
            value=valor, year=year, index=3, df_by_year=df_by_year
        )
        for valor in valores_columna
    ]

    # Sum the previous radicado in previous years
    data_frame[historical_df.columns[8]] = data_frame[historical_df.columns[8]].astype(
        int
    ) + data_frame["is_previous_radicado"].fillna(0).astype(int)

    data_frame[historical_df.columns[9]] = data_frame[historical_df.columns[9]].astype(
        int
    ) + data_frame["is_previous_key"].fillna(0).astype(int)

    # Ignore no needed temp columns
    data_frame = data_frame.iloc[:, :13]
    # print(data_frame)  # <-- this is the final data frame
    # Set up unique columns
    data_frame.columns = historical_df.columns
    return data_frame


def add_number_of_duplicates(
    data_frame: pd.DataFrame,
    historical_df: pd.DataFrame,
    index_col: int,
    historical_index_col: int,
) -> None:
    ## Key column: (radicado number + concepto value)
    key_col = data_frame.columns[index_col]
    key_count = data_frame[key_col].map(data_frame[key_col].value_counts())
    data_frame[historical_df.columns[historical_index_col]] = key_count


def validate_values(acm_file: str) -> None:
    try:
        # Extract data from propuesta de pagos file
        propuesta_pago_df = values_validation.read_excel(
            values_validation.file_path, values_validation.sheet_name
        )
        # Get previous data frame
        historical_df: pd.DataFrame = values_validation.read_excel(
            values_validation.previous_file, values_validation.sheet_name
        )
        # Get the ACM report data fame
        acm_report: pd.DataFrame = values_validation.read_excel(
            acm_file, "FCT_RS_REPORTE_WS_AUDITORIA"
        )
        # Delete te no needed columns and rows
        acm_report = acm_report.iloc[3:, 1:]
        acm_report.columns = acm_report.iloc[0]
        acm_report = acm_report.iloc[1:].reset_index(drop=True)
        acm_report = acm_report[["id cuenta", "valor aprobado", "Valor Liquidado"]]

        # Validate and extract the important data from propuesta and acm report
        propuesta_df: pd.DataFrame = extract_data_from_propuesta(propuesta_pago_df)
        merged_df: pd.DataFrame = cross_file(propuesta_df, acm_report)
        # Apply formulas to validate inconsistencies
        filled_df: pd.DataFrame = apply_formulas(merged_df, historical_df)
        # Report inconsistencies
        report_inconsistencies(filled_df)
        # Save the final file into temp file folder
        filled_df.to_excel(
            values_validation.temp_file,
            sheet_name=values_validation.sheet_name,
            index=False,
        )
        return True, f"Function '{validate_values.__name__}' executed successfully"

    except Exception as e:
        return False, f"Error: {e}"


def report_inconsistencies(data_frame: pd.DataFrame) -> None:
    """Method to generate a report of inconsistencies and save it into tbe correct file"""
    try:
        # 1. Values validation
        values_validation.save_inconsistencies(
            data_frame=data_frame,
            exception_sheet_name="VALIDACION VALORES",
            exception_col_idx=0,
            validation_col_idx=7,
            col_idx_to_except=3,
            inconsistencies_sheet_name="ValidacionValores",
        )
        # 2. Radicados number duplicated
        values_validation.save_inconsistencies_values(
            data_frame=data_frame,
            exception_sheet_name="VALIDACION DUPLICADOS",
            exception_col=0,
            validation_col=8,
            list_col=1,
            inconsistencies_sheet_name="ValidacionRadicadosDuplicados",
        )
        # 3. Key duplicated
        values_validation.save_inconsistencies_values(
            data_frame=data_frame,
            exception_sheet_name="VALIDACION DUPLICADOS",
            exception_col=1,
            validation_col=9,
            list_col=3,
            inconsistencies_sheet_name="ValidacionKeyDuplicados",
        )
        # 4. Radicado format
        values_validation.save_inconsistencies(
            data_frame=data_frame,
            exception_sheet_name="VALIDACION FORMATOS",
            exception_col_idx=0,
            validation_col_idx=10,
            col_idx_to_except=3,
            inconsistencies_sheet_name="ValidacionRadicadoFormato",
        )
        # 5. Valor 100% format
        values_validation.save_inconsistencies(
            data_frame=data_frame,
            exception_sheet_name="VALIDACION FORMATOS",
            exception_col_idx=1,
            validation_col_idx=11,
            col_idx_to_except=2,
            inconsistencies_sheet_name="ValidacionValor100Formato",
        )
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False


# Call the main function
def main(params: dict) -> bool:
    try:
        global values_validation
        if values_validation is None:
            values_validation = ValuesValidation(
                file_path=params.get("file_path"),
                inconsistencies_file=params.get("inconsistencies_file"),
                exception_file=params.get("exception_file"),
                sheet_name=params.get("sheet_name"),
                file_name=params.get("file_name"),
                previous_file=params.get("previous_file"),
                temp_file=params.get("temp_file"),
                historic_file=params.get("historic_file"),
            )
            return True
    except Exception as e:
        print(f"Error: {e}")
        return False


if __name__ == "__main__":
    params = {
        "file_path": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Input\PROPUESTA DE PAGO 1 Y 2  (02-01-2025).xlsx",
        "inconsistencies_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Temp\InconsistenciasBasePagosRedAsistencial.xlsx",
        "exception_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Input\EXCEPCIONES BASE PAGOS RED ASISTENCIAL.xlsx",
        "sheet_name": "Propuesta",
        "file_name": "PROPUESTA DE PAGO 1 Y 2  (02-01-2025).xlsx",
        "previous_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Output\Validacion valores\Validacion valores 04122024.xlsx",
        "temp_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Temp\Pagos red asistencial 18122024.xlsx",
        "historic_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Input\VALIDADOR PAGOS.xlsx",
    }
    print(main(params))
    incomes = r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Temp\FCT_RS_REPORTE_WS_AUDITORIA.xlsx"
    print(validate_values(incomes))
