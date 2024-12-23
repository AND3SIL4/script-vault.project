import pandas as pd  # type: ignore
from openpyxl import load_workbook  # type: ignore
from datetime import datetime


def update_validador_pagos_file(params: dict) -> tuple:
    try:
        # Get the variables from the params dictionary
        validador_pagos_file = params.get("validador_pagos_file")
        temp_file = params.get("temp_file")

        # Set local variables
        year = datetime.now().year

        # Validate if all the required inputs are present
        if not all([validador_pagos_file, temp_file]):
            raise Exception("A required input is missing, please check and try again")

        # Read the validador_pagos_file using pandas
        temp_df: pd.DataFrame = pd.read_excel(temp_file, "Propuesta", engine="openpyxl")

        # Get only the needed columns from data frame
        temp_df = temp_df.iloc[:, :4]

        # Load workbook
        book = load_workbook(validador_pagos_file)
        sheet_names = book.sheetnames

        # Check if the sheet for the current year exists
        if str(year) in sheet_names:
            sheet = book[str(year)]
            # Get the max row number to append data without overwriting
            start_row = sheet.max_row + 1
        else:
            # If sheet doesn't exist, create it
            sheet = book.create_sheet(str(year))
            start_row = 1

        # Append the new data to the sheet
        for row in temp_df.itertuples(index=False, name=None):
            sheet.append(row)

        # Save the workbook with the changes
        book.save(validador_pagos_file)

        return (
            True,
            f"Function '{update_validador_pagos_file.__name__}' executed successfully",
        )

    except Exception as e:
        return (False, f"Error: {e}")


def update_final_file(params: dict) -> tuple:
    try:
        # Get the variables from the params dictionary
        historical_file = params.get("historical_file")
        values_validation_file = params.get("values_validation_file")
        final_path = params.get("final_path")

        # Validate if all the required inputs are present
        if not all([values_validation_file, historical_file]):
            raise Exception("A required input is missing, please check and try again")

        # Create data frames
        historical_df: pd.DataFrame = pd.read_excel(
            historical_file, sheet_name="Propuesta", engine="openpyxl"
        )
        values_validation_df: pd.DataFrame = pd.read_excel(
            values_validation_file, engine="openpyxl", sheet_name="Propuesta"
        )

        # Concat data frames
        final_df: pd.DataFrame = pd.concat(
            [historical_df, values_validation_df], ignore_index=True
        )

        # Save the final file
        final_df.to_excel(
            final_path,
            sheet_name="Propuesta",
            engine="openpyxl",
            index=False,
        )

        return (
            True,
            f"Function '{update_final_file.__name__}' executed successfully, final file saved correctly",
        )

    except Exception as e:
        return (False, f"Error: {e}")


if __name__ == "__main__":
    # params = {
    #     "validador_pagos_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Input\VALIDADOR PAGOS - copia.xlsx",
    #     "temp_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Temp\Pagos red asistencial 18122024.xlsx",
    # }

    params = {
        "values_validation_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Temp\Pagos red asistencial 18122024.xlsx",
        "historical_file": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Output\Validacion valores\Validacion valores 23102024.xlsx",
        "final_path": r"C:\ProgramData\AutomationAnywhere\Bots\AD_GI_BasePagosRedAsistencial_SabanaPagosBasesSiniestralidad\Output\Validacion valores\Validacion valores 23122024.xlsx",
    }

    print(update_final_file(params))
